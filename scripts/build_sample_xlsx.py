"""Convert docs/samples/customers_sample.csv to a polished .xlsx workbook.

Run this once whenever the CSV changes:

    python scripts/build_sample_xlsx.py

The output is committed alongside the CSV so users can download either format
straight from the repo.
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
CSV_PATH = ROOT / "docs" / "samples" / "customers_sample.csv"
XLSX_PATH = ROOT / "docs" / "samples" / "customers_sample.xlsx"


def main() -> int:
    if not CSV_PATH.exists():
        print(f"Missing {CSV_PATH}", file=sys.stderr)
        return 1

    rows = list(csv.reader(CSV_PATH.read_text(encoding="utf-8").splitlines()))
    if not rows:
        print("Empty CSV", file=sys.stderr)
        return 1

    header, *body = rows

    wb = Workbook()
    ws = wb.active
    ws.title = "customers"

    header_fill = PatternFill("solid", fgColor="0F2740")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    depot_fill = PatternFill("solid", fgColor="FFF4D6")
    center = Alignment(horizontal="center", vertical="center")

    ws.append(header)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for row in body:
        ws.append(row)

    # Highlight depot row (demand = 0).
    demand_idx = header.index("demand") + 1
    for excel_row in ws.iter_rows(min_row=2, values_only=False):
        try:
            demand_value = int(float(excel_row[demand_idx - 1].value or 0))
        except (TypeError, ValueError):
            demand_value = 1
        if demand_value == 0:
            for cell in excel_row:
                cell.fill = depot_fill
                cell.font = Font(bold=True)
            break

    widths = [22, 32, 12, 12, 10, 8, 8, 10]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"

    wb.save(XLSX_PATH)
    print(f"Wrote {XLSX_PATH.relative_to(ROOT)} ({len(body)} rows)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
